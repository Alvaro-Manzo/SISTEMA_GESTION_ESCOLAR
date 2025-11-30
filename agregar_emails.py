"""
Script auxiliar para agregar emails a los estudiantes
"""

import openpyxl

def agregar_emails_estudiantes():
    """Agrega emails a los estudiantes del Excel"""
    
    print("\n📧 AGREGAR EMAILS A ESTUDIANTES")
    print("="*60)
    
    try:
        archivo = 'grupo001.xlsx'
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Verificar/agregar columna EMAIL
        headers = [cell.value for cell in ws[1]]
        if 'EMAIL' not in headers:
            ws.cell(1, 5).value = 'EMAIL'
            print("✅ Columna EMAIL creada")
        
        idx_email = 5  # Columna E
        
        print("\n📋 Lista de estudiantes:")
        print("-"*60)
        
        estudiantes_sin_email = []
        
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, 1).value
            email_actual = ws.cell(row, idx_email).value
            
            if nombre:
                if not email_actual or str(email_actual).strip() == '':
                    estudiantes_sin_email.append((row, nombre))
                    print(f"{row-1}. {nombre:<35} - ⊘ Sin email")
                else:
                    print(f"{row-1}. {nombre:<35} - ✅ {email_actual}")
        
        print("\n" + "="*60)
        print(f"Total de estudiantes sin email: {len(estudiantes_sin_email)}")
        print("="*60)
        
        if not estudiantes_sin_email:
            print("\n✅ Todos los estudiantes ya tienen email registrado!")
            return
        
        print("\n📝 Opciones:")
        print("1. Agregar emails uno por uno")
        print("2. Usar formato estándar (nombre@dominio.com)")
        print("3. Cancelar")
        
        opcion = input("\nSelecciona una opción (1-3): ").strip()
        
        if opcion == '1':
            print("\n💡 Ingresa el email de cada estudiante (o ENTER para saltar)")
            print("-"*60)
            
            for row, nombre in estudiantes_sin_email:
                email = input(f"\nEmail para {nombre}: ").strip()
                if email and '@' in email:
                    ws.cell(row, idx_email).value = email
                    print(f"   ✅ Email agregado")
                elif email:
                    print(f"   ⚠️  Email inválido, se omitió")
            
            wb.save(archivo)
            print(f"\n✅ Cambios guardados en {archivo}")
        
        elif opcion == '2':
            dominio = input("\n📧 Ingresa el dominio (ej. estudiantes.edu.mx): ").strip()
            
            if not dominio:
                print("❌ Operación cancelada")
                return
            
            print(f"\n⏳ Generando emails con formato nombre@{dominio}...")
            
            for row, nombre in estudiantes_sin_email:
                # Convertir nombre a email
                # JUAN PEREZ -> juan.perez@dominio
                nombre_email = nombre.lower().replace(' ', '.').replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u').replace('ñ', 'n')
                email = f"{nombre_email}@{dominio}"
                
                ws.cell(row, idx_email).value = email
                print(f"   {nombre:<35} → {email}")
            
            wb.save(archivo)
            print(f"\n✅ Emails generados y guardados en {archivo}")
        
        else:
            print("\n❌ Operación cancelada")
    
    except Exception as e:
        print(f"\n❌ Error: {e}")


def ver_emails():
    """Muestra todos los estudiantes con sus emails"""
    try:
        archivo = 'grupo001.xlsx'
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        if 'EMAIL' not in headers:
            print("\n⚠️  No hay columna EMAIL en el archivo")
            print("   Ejecuta la opción 1 del menú principal primero")
            return
        
        idx_email = headers.index('EMAIL')
        
        print("\n📧 EMAILS REGISTRADOS")
        print("="*60)
        print(f"\n{'NOMBRE':<35} {'EMAIL':<40}")
        print("-"*60)
        
        total = 0
        con_email = 0
        
        for row in range(2, ws.max_row + 1):
            nombre = ws.cell(row, 1).value
            email = ws.cell(row, idx_email + 1).value
            
            if nombre:
                total += 1
                if email and str(email).strip() != '' and '@' in str(email):
                    print(f"{nombre:<35} {email:<40}")
                    con_email += 1
                else:
                    print(f"{nombre:<35} {'⊘ Sin email':<40}")
        
        print("-"*60)
        print(f"\nTotal: {total} estudiantes")
        print(f"Con email: {con_email} ({con_email/total*100:.1f}%)")
        print(f"Sin email: {total-con_email}")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")


def menu():
    """Menú del gestor de emails"""
    while True:
        print("\n" + "="*60)
        print("   📧 GESTOR DE EMAILS PARA ESTUDIANTES 📧")
        print("="*60)
        print("\n1. ➕ Agregar emails a estudiantes")
        print("2. 👀 Ver emails registrados")
        print("3. 🔙 Salir")
        print("\n" + "="*60)
        
        opcion = input("\nSelecciona una opción (1-3): ").strip()
        
        if opcion == '1':
            agregar_emails_estudiantes()
        elif opcion == '2':
            ver_emails()
        elif opcion == '3':
            print("\n👋 ¡Hasta pronto!")
            break
        else:
            print("\n❌ Opción inválida")
        
        input("\n⏎ Presiona ENTER para continuar...")


if __name__ == "__main__":
    menu()

import time
import openpyxl
from openpyxl.styles import Font, Pattdef mostrar_menu_estudiante():
    """Muestra el menú del estudiante"""
    print("\n" + "="*60)
    print("   👨‍🎓 PANEL DE ESTUDIANTE 👨‍🎓")
    print("="*60)
    print("\n1. 🔍 Consultar mi calificación (con número de cuenta)")
    print("2. 📊 Ver estadísticas del grupo")
    print("3. 🔙 Volver al menú principal")
    print("\n" + "="*60)
    print("\n💡 Necesitas tu NÚMERO DE CUENTA para consultar tu calificación")Alignment
import os
import getpass

# Contraseña del administrador (maestro)
ADMIN_PASSWORD = "admin123"  # Cambia esta contraseña por la que desees

def limpiar_pantalla():
    """Limpia la consola"""
    os.system('clear' if os.name != 'nt' else 'cls')

def cargar_datos_excel(archivo='grupo001.xlsx'):
    """Carga los datos del archivo Excel"""
    try:
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        estudiantes = {}
        
        # Leer datos (empezando desde la fila 2 para saltar encabezados)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Si hay nombre
                nombre = str(row[0]).strip().upper()
                calificacion = row[1] if row[1] is not None else 0
                # Verificar si existe número de cuenta (columna 4)
                numero_cuenta = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                estudiantes[nombre] = {
                    'calificacion': calificacion,
                    'estado': 'APROBADO' if calificacion >= 6 else 'REPROBADO',
                    'numero_cuenta': numero_cuenta
                }
        
        return estudiantes, wb, ws
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {archivo}")
        return None, None, None
    except Exception as e:
        print(f"❌ Error al cargar el archivo: {e}")
        return None, None, None

def mostrar_menu():
    """Muestra el menú principal de acceso"""
    print("\n" + "="*60)
    print("   📚 SISTEMA DE CALIFICACIONES 📚")
    print("="*60)
    print("\n1. 👨‍🎓 Acceso para ESTUDIANTES")
    print("2. �‍🏫 Acceso para ADMINISTRADOR (Maestro)")
    print("3. 🚪 Salir")
    print("\n" + "="*60)

def mostrar_menu_admin():
    """Muestra el menú del administrador"""
    print("\n" + "="*60)
    print("   👨‍🏫 PANEL DE ADMINISTRADOR 👨‍🏫")
    print("="*60)
    print("\n1. �🔍 Consultar calificación de un estudiante")
    print("2. 📊 Ver todos los estudiantes")
    print("3. ➕ Agregar nuevo estudiante")
    print("4. ✏️  Modificar calificación")
    print("5. 🗑️  Eliminar estudiante")
    print("6. 📈 Estadísticas del grupo")
    print("7. 🔙 Volver al menú principal")
    print("\n" + "="*60)

def mostrar_menu_estudiante():
    """Muestra el menú del estudiante"""
    print("\n" + "="*60)
    print("   👨‍🎓 PANEL DE ESTUDIANTE 👨‍🎓")
    print("="*60)
    print("\n1. 🔍 Consultar mi calificación")
    print("2. � Ver estadísticas del grupo")
    print("3. 🔙 Volver al menú principal")
    print("\n" + "="*60)

def consultar_estudiante(estudiantes, modo='admin'):
    """Consulta la calificación de un estudiante específico"""
    print("\n" + "-"*60)
    if modo == 'admin':
        print("   🔍 CONSULTAR CALIFICACIÓN DE ESTUDIANTE")
    else:
        print("   🔍 CONSULTAR MI CALIFICACIÓN")
    print("-"*60)
    
    if modo == 'estudiante':
        # Los estudiantes usan número de cuenta
        numero_cuenta = input("\nIngresa tu NÚMERO DE CUENTA (ej. 324012345): ").strip()
        
        # Buscar estudiante por número de cuenta
        estudiante_encontrado = None
        nombre_estudiante = None
        
        for nombre, info in estudiantes.items():
            if info.get('numero_cuenta') == numero_cuenta:
                estudiante_encontrado = info
                nombre_estudiante = nombre
                break
        
        if estudiante_encontrado:
            print(f"\n✅ ¡Bienvenido!")
            print(f"\n   👤 Nombre: {nombre_estudiante}")
            print(f"   � Número de cuenta: {numero_cuenta}")
            print(f"   �📝 Calificación: {estudiante_encontrado['calificacion']}")
            print(f"   📊 Estado: {estudiante_encontrado['estado']}")
            
            if estudiante_encontrado['estado'] == 'APROBADO':
                print("\n   🎉 ¡Felicidades! Has aprobado la materia.")
            else:
                print("\n   😔 Lo siento, no has aprobado. ¡Sigue esforzándote!")
        else:
            print(f"\n❌ Número de cuenta '{numero_cuenta}' no encontrado.")
            print("   💡 Verifica que ingresaste correctamente tu número de cuenta.")
            print("   💡 Si olvidaste tu número, contacta a tu maestro.")
    else:
        # Admin usa nombre
        nombre = input("\nIngresa el nombre completo en MAYÚSCULAS (ej. JUAN PEREZ): ").strip().upper()
        
        if nombre in estudiantes:
            info = estudiantes[nombre]
            print(f"\n✅ Estudiante encontrado!")
            print(f"\n   👤 Nombre: {nombre}")
            print(f"   � Número de cuenta: {info.get('numero_cuenta', 'No asignado')}")
            print(f"   📝 Calificación: {info['calificacion']}")
            print(f"   📊 Estado: {info['estado']}")
            
            if info['estado'] == 'APROBADO':
                print("\n   🎉 ¡Felicidades! Has aprobado la materia.")
            else:
                print("\n   😔 Lo siento, no has aprobado. ¡Sigue esforzándote!")
        else:
            print(f"\n❌ El estudiante '{nombre}' no se encuentra en el sistema.")

def ver_todos_estudiantes(estudiantes):
    """Muestra todos los estudiantes y sus calificaciones"""
    print("\n" + "-"*60)
    print("   📊 LISTA COMPLETA DE ESTUDIANTES")
    print("-"*60)
    
    if not estudiantes:
        print("\n❌ No hay estudiantes registrados.")
        return
    
    # Ordenar por nombre
    estudiantes_ordenados = sorted(estudiantes.items())
    
    print(f"\n{'NOMBRE':<30} {'CALIFICACIÓN':<15} {'ESTADO':<15}")
    print("-"*60)
    
    for nombre, info in estudiantes_ordenados:
        simbolo = "✅" if info['estado'] == 'APROBADO' else "❌"
        print(f"{nombre:<30} {info['calificacion']:<15} {simbolo} {info['estado']:<15}")
    
    print("-"*60)
    print(f"Total de estudiantes: {len(estudiantes)}")

def agregar_estudiante(estudiantes, wb, ws):
    """Agrega un nuevo estudiante al sistema"""
    print("\n" + "-"*60)
    print("   ➕ AGREGAR NUEVO ESTUDIANTE")
    print("-"*60)
    
    nombre = input("\nIngresa el nombre completo en MAYÚSCULAS: ").strip().upper()
    
    if nombre in estudiantes:
        print(f"\n⚠️  El estudiante '{nombre}' ya existe en el sistema.")
        return estudiantes
    
    try:
        calificacion = float(input("Ingresa la calificación (0-10): "))
        if calificacion < 0 or calificacion > 10:
            print("\n❌ La calificación debe estar entre 0 y 10.")
            return estudiantes
    except ValueError:
        print("\n❌ Calificación inválida.")
        return estudiantes
    
    # Generar número de cuenta único
    import random
    while True:
        numero_cuenta = f"3240{random.randint(10000, 99999)}"
        # Verificar que no exista
        existe = False
        for info in estudiantes.values():
            if info.get('numero_cuenta') == numero_cuenta:
                existe = True
                break
        if not existe:
            break
    
    # Agregar al diccionario
    estado = 'APROBADO' if calificacion >= 6 else 'REPROBADO'
    estudiantes[nombre] = {
        'calificacion': calificacion,
        'estado': estado,
        'numero_cuenta': numero_cuenta
    }
    
    # Agregar al Excel
    nueva_fila = ws.max_row + 1
    ws[f'A{nueva_fila}'] = nombre
    ws[f'B{nueva_fila}'] = calificacion
    ws[f'C{nueva_fila}'] = f'=IF(B{nueva_fila}>=6, "Aprobado", "Reprobado")'
    ws[f'D{nueva_fila}'] = numero_cuenta
    
    try:
        wb.save('grupo001.xlsx')
        print(f"\n✅ Estudiante '{nombre}' agregado exitosamente!")
        print(f"   📝 Calificación: {calificacion}")
        print(f"   📊 Estado: {estado}")
        print(f"   🔐 Número de cuenta: {numero_cuenta}")
        print(f"\n   ⚠️  IMPORTANTE: Proporciona este número de cuenta al estudiante")
    except Exception as e:
        print(f"\n❌ Error al guardar: {e}")
    
    return estudiantes

def modificar_calificacion(estudiantes, wb, ws):
    """Modifica la calificación de un estudiante"""
    print("\n" + "-"*60)
    print("   ✏️  MODIFICAR CALIFICACIÓN")
    print("-"*60)
    
    nombre = input("\nIngresa el nombre del estudiante: ").strip().upper()
    
    if nombre not in estudiantes:
        print(f"\n❌ El estudiante '{nombre}' no se encuentra en el sistema.")
        return estudiantes
    
    print(f"\nCalificación actual: {estudiantes[nombre]['calificacion']}")
    
    try:
        nueva_calificacion = float(input("Ingresa la nueva calificación (0-10): "))
        if nueva_calificacion < 0 or nueva_calificacion > 10:
            print("\n❌ La calificación debe estar entre 0 y 10.")
            return estudiantes
    except ValueError:
        print("\n❌ Calificación inválida.")
        return estudiantes
    
    # Actualizar diccionario
    estado = 'APROBADO' if nueva_calificacion >= 6 else 'REPROBADO'
    estudiantes[nombre]['calificacion'] = nueva_calificacion
    estudiantes[nombre]['estado'] = estado
    
    # Actualizar Excel
    for row in range(2, ws.max_row + 1):
        if str(ws[f'A{row}'].value).strip().upper() == nombre:
            ws[f'B{row}'] = nueva_calificacion
            break
    
    try:
        wb.save('grupo001.xlsx')
        print(f"\n✅ Calificación actualizada exitosamente!")
        print(f"   📝 Nueva calificación: {nueva_calificacion}")
        print(f"   📊 Estado: {estado}")
    except Exception as e:
        print(f"\n❌ Error al guardar: {e}")
    
    return estudiantes

def mostrar_estadisticas(estudiantes):
    """Muestra estadísticas del grupo"""
    print("\n" + "-"*60)
    print("   📈 ESTADÍSTICAS DEL GRUPO")
    print("-"*60)
    
    if not estudiantes:
        print("\n❌ No hay estudiantes registrados.")
        return
    
    calificaciones = [info['calificacion'] for info in estudiantes.values()]
    aprobados = sum(1 for info in estudiantes.values() if info['estado'] == 'APROBADO')
    reprobados = len(estudiantes) - aprobados
    
    promedio = sum(calificaciones) / len(calificaciones)
    maxima = max(calificaciones)
    minima = min(calificaciones)
    
    print(f"\n📊 Total de estudiantes: {len(estudiantes)}")
    print(f"✅ Aprobados: {aprobados} ({aprobados/len(estudiantes)*100:.1f}%)")
    print(f"❌ Reprobados: {reprobados} ({reprobados/len(estudiantes)*100:.1f}%)")
    print(f"\n📈 Calificación promedio: {promedio:.2f}")
    print(f"🏆 Calificación más alta: {maxima}")
    print(f"📉 Calificación más baja: {minima}")
    
    # Estudiante(s) con mejor calificación
    mejores = [nombre for nombre, info in estudiantes.items() if info['calificacion'] == maxima]
    print(f"\n🥇 Mejor(es) estudiante(s):")
    for nombre in mejores:
        print(f"   - {nombre}")

def eliminar_estudiante(estudiantes, wb, ws):
    """Elimina un estudiante del sistema"""
    print("\n" + "-"*60)
    print("   🗑️  ELIMINAR ESTUDIANTE")
    print("-"*60)
    
    nombre = input("\nIngresa el nombre del estudiante a eliminar: ").strip().upper()
    
    if nombre not in estudiantes:
        print(f"\n❌ El estudiante '{nombre}' no se encuentra en el sistema.")
        return estudiantes
    
    print(f"\n⚠️  ¿Estás seguro de eliminar a '{nombre}'?")
    confirmacion = input("Escribe 'SI' para confirmar: ").strip().upper()
    
    if confirmacion != 'SI':
        print("\n❌ Operación cancelada.")
        return estudiantes
    
    # Eliminar del diccionario
    del estudiantes[nombre]
    
    # Eliminar del Excel
    fila_eliminar = None
    for row in range(2, ws.max_row + 1):
        if str(ws[f'A{row}'].value).strip().upper() == nombre:
            fila_eliminar = row
            break
    
    if fila_eliminar:
        ws.delete_rows(fila_eliminar, 1)
        try:
            wb.save('grupo001.xlsx')
            print(f"\n✅ Estudiante '{nombre}' eliminado exitosamente.")
        except Exception as e:
            print(f"\n❌ Error al guardar: {e}")
    
    return estudiantes

def verificar_admin():
    """Verifica la contraseña del administrador"""
    print("\n" + "-"*60)
    print("   🔐 ACCESO DE ADMINISTRADOR")
    print("-"*60)
    intentos = 3
    
    while intentos > 0:
        password = getpass.getpass(f"\nIngresa la contraseña (intentos restantes: {intentos}): ")
        
        if password == ADMIN_PASSWORD:
            print("\n✅ Acceso concedido. Bienvenido, Maestro.")
            time.sleep(1)
            return True
        else:
            intentos -= 1
            if intentos > 0:
                print(f"❌ Contraseña incorrecta. Te quedan {intentos} intentos.")
            else:
                print("\n❌ Acceso denegado. Demasiados intentos fallidos.")
                time.sleep(2)
    
    return False

def panel_estudiante(estudiantes):
    """Panel de acceso para estudiantes"""
    while True:
        mostrar_menu_estudiante()
        opcion = input("\nSelecciona una opción (1-3): ").strip()
        
        if opcion == '1':
            consultar_estudiante(estudiantes, modo='estudiante')
        elif opcion == '2':
            mostrar_estadisticas(estudiantes)
        elif opcion == '3':
            print("\n🔙 Volviendo al menú principal...")
            time.sleep(1)
            break
        else:
            print("\n❌ Opción inválida. Por favor selecciona 1-3.")
        
        input("\n⏎ Presiona ENTER para continuar...")
        limpiar_pantalla()

def panel_admin(estudiantes, wb, ws):
    """Panel de acceso para administradores"""
    if not verificar_admin():
        return estudiantes
    
    limpiar_pantalla()
    
    while True:
        mostrar_menu_admin()
        opcion = input("\nSelecciona una opción (1-7): ").strip()
        
        if opcion == '1':
            consultar_estudiante(estudiantes, modo='admin')
        elif opcion == '2':
            ver_todos_estudiantes(estudiantes)
        elif opcion == '3':
            estudiantes = agregar_estudiante(estudiantes, wb, ws)
        elif opcion == '4':
            estudiantes = modificar_calificacion(estudiantes, wb, ws)
        elif opcion == '5':
            estudiantes = eliminar_estudiante(estudiantes, wb, ws)
        elif opcion == '6':
            mostrar_estadisticas(estudiantes)
        elif opcion == '7':
            print("\n🔙 Volviendo al menú principal...")
            time.sleep(1)
            break
        else:
            print("\n❌ Opción inválida. Por favor selecciona 1-7.")
        
        input("\n⏎ Presiona ENTER para continuar...")
        limpiar_pantalla()
    
    return estudiantes

def main():
    """Función principal del sistema"""
    limpiar_pantalla()
    print("\n" + "="*60)
    print("   🎓 SISTEMA DE CALIFICACIONES 🎓")
    print("="*60)
    time.sleep(1)
    
    # Cargar datos
    print("\n⏳ Cargando datos...")
    estudiantes, wb, ws = cargar_datos_excel()
    
    if estudiantes is None:
        print("\n❌ No se pudo iniciar el sistema.")
        return
    
    print(f"✅ Datos cargados: {len(estudiantes)} estudiantes encontrados.")
    time.sleep(1)
    limpiar_pantalla()
    
    while True:
        mostrar_menu()
        opcion = input("\nSelecciona una opción (1-3): ").strip()
        
        if opcion == '1':
            limpiar_pantalla()
            print("\n👨‍🎓 Bienvenido, Estudiante")
            time.sleep(1)
            limpiar_pantalla()
            panel_estudiante(estudiantes)
            limpiar_pantalla()
        elif opcion == '2':
            limpiar_pantalla()
            estudiantes = panel_admin(estudiantes, wb, ws)
            limpiar_pantalla()
        elif opcion == '3':
            print("\n👋 ¡Gracias por usar el sistema! Hasta pronto.")
            time.sleep(1)
            break
        else:
            print("\n❌ Opción inválida. Por favor selecciona 1-3.")
            time.sleep(1)
            limpiar_pantalla()

if __name__ == "__main__":
    main()

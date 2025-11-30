"""
Script para generar números de cuenta únicos para los estudiantes
Formato: 3240XXXXX (donde XXXXX son números aleatorios únicos)
"""

import openpyxl
import random

def generar_numero_cuenta():
    """Genera un número de cuenta único con formato 3240XXXXX"""
    sufijo = random.randint(10000, 99999)
    return f"3240{sufijo}"

def agregar_numeros_cuenta(archivo_excel='grupo001.xlsx'):
    """Agrega números de cuenta únicos a todos los estudiantes"""
    try:
        # Cargar archivo
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
        
        # Verificar si ya existe la columna de número de cuenta
        if ws.cell(1, 4).value != 'NUMERO DE CUENTA':
            # Agregar encabezado
            ws.cell(1, 4).value = 'NUMERO DE CUENTA'
            
            # Generar números únicos
            numeros_generados = set()
            
            # Recorrer todas las filas con estudiantes
            for row in range(2, ws.max_row + 1):
                nombre = ws.cell(row, 1).value
                if nombre:  # Si hay un nombre
                    # Generar número único
                    while True:
                        numero_cuenta = generar_numero_cuenta()
                        if numero_cuenta not in numeros_generados:
                            numeros_generados.add(numero_cuenta)
                            break
                    
                    # Asignar número de cuenta
                    ws.cell(row, 4).value = numero_cuenta
                    print(f"✅ {nombre:<30} → {numero_cuenta}")
            
            # Guardar cambios
            wb.save(archivo_excel)
            print(f"\n✅ Números de cuenta generados exitosamente!")
            print(f"📁 Archivo actualizado: {archivo_excel}")
            print(f"📊 Total de cuentas generadas: {len(numeros_generados)}")
            
            # Crear archivo con las credenciales
            crear_archivo_credenciales(ws, numeros_generados)
            
        else:
            print("⚠️  Los números de cuenta ya existen en el archivo.")
            print("💡 Si deseas regenerarlos, elimina la columna D primero.")
    
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {archivo_excel}")
    except Exception as e:
        print(f"❌ Error: {e}")

def crear_archivo_credenciales(ws, numeros_generados):
    """Crea un archivo de texto con todas las credenciales"""
    try:
        with open('credenciales_estudiantes.txt', 'w', encoding='utf-8') as f:
            f.write("="*70 + "\n")
            f.write("  CREDENCIALES DE ACCESO - SISTEMA DE CALIFICACIONES\n")
            f.write("="*70 + "\n\n")
            f.write("⚠️  CONFIDENCIAL - Distribuir a cada estudiante SOLO su número de cuenta\n\n")
            f.write("-"*70 + "\n")
            f.write(f"{'NOMBRE DEL ESTUDIANTE':<35} {'NÚMERO DE CUENTA':<20}\n")
            f.write("-"*70 + "\n")
            
            for row in range(2, ws.max_row + 1):
                nombre = ws.cell(row, 1).value
                numero_cuenta = ws.cell(row, 4).value
                if nombre and numero_cuenta:
                    f.write(f"{nombre:<35} {numero_cuenta:<20}\n")
            
            f.write("-"*70 + "\n")
            f.write(f"\nTotal de estudiantes: {len(numeros_generados)}\n")
            f.write("\n⚠️  IMPORTANTE: Cada estudiante debe conocer SOLO su propio número de cuenta\n")
            f.write("para garantizar la privacidad de las calificaciones.\n")
        
        print(f"\n📄 Archivo de credenciales creado: credenciales_estudiantes.txt")
        print("⚠️  IMPORTANTE: Este archivo contiene información confidencial")
    
    except Exception as e:
        print(f"❌ Error al crear archivo de credenciales: {e}")

if __name__ == "__main__":
    print("\n🔐 GENERADOR DE NÚMEROS DE CUENTA\n")
    print("Este script agregará números de cuenta únicos a todos los estudiantes")
    print("Formato: 3240XXXXX (donde XXXXX son números aleatorios)\n")
    
    confirmacion = input("¿Deseas continuar? (SI/NO): ").strip().upper()
    
    if confirmacion == 'SI':
        agregar_numeros_cuenta()
    else:
        print("\n❌ Operación cancelada.")

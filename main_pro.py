"""
Sistema de Calificaciones Académico - Versión PRO 2.0
Autor: Sistema Educativo
Fecha: 2025

Características Pro:
- Sistema de configuración JSON
- Logging avanzado de operaciones
- Backups automáticos
- Interfaz con colores
- Exportación de reportes
- Validaciones robustas
- Caché de datos
"""

import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import sys
import json
import logging
from datetime import datetime
import shutil
import csv
from pathlib import Path
import getpass
from typing import Dict, Tuple, Optional

# Intentar importar colorama para colores en terminal
try:
    from colorama import init, Fore, Back, Style
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False
    # Fallback si no está instalado
    class Fore:
        RED = GREEN = YELLOW = BLUE = MAGENTA = CYAN = WHITE = RESET = ""
    class Back:
        RED = GREEN = YELLOW = BLUE = MAGENTA = CYAN = WHITE = RESET = ""
    class Style:
        BRIGHT = DIM = NORMAL = RESET_ALL = ""


class ConfigManager:
    """Gestor de configuración del sistema"""
    
    def __init__(self, config_file='config.json'):
        self.config_file = config_file
        self.config = self.cargar_config()
    
    def cargar_config(self) -> dict:
        """Carga la configuración desde el archivo JSON"""
        try:
            with open(self.config_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            logging.warning(f"Archivo de configuración {self.config_file} no encontrado. Usando valores por defecto.")
            return self.config_por_defecto()
        except json.JSONDecodeError as e:
            logging.error(f"Error al leer configuración: {e}")
            return self.config_por_defecto()
    
    def config_por_defecto(self) -> dict:
        """Retorna configuración por defecto"""
        return {
            "seguridad": {"admin_password": "admin123", "max_intentos_login": 3},
            "archivos": {"excel_principal": "grupo001.xlsx"},
            "calificaciones": {"minima_aprobatoria": 6.0},
            "interfaz": {"usar_colores": True, "limpiar_pantalla": True},
            "backups": {"automaticos": True, "max_backups": 10},
            "logs": {"activar": True, "nivel": "INFO"}
        }
    
    def get(self, *keys, default=None):
        """Obtiene un valor de configuración usando path de keys"""
        value = self.config
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return default
        return value


class LogManager:
    """Gestor de logs del sistema"""
    
    def __init__(self, config: ConfigManager):
        self.config = config
        self.log_dir = Path(config.get('archivos', 'directorio_logs', default='logs'))
        self.setup_logging()
    
    def setup_logging(self):
        """Configura el sistema de logging"""
        if not self.config.get('logs', 'activar', default=True):
            return
        
        # Crear directorio de logs si no existe
        self.log_dir.mkdir(exist_ok=True)
        
        # Archivo de log con fecha
        log_file = self.log_dir / f"sistema_{datetime.now().strftime('%Y%m%d')}.log"
        
        # Configurar logging
        nivel = self.config.get('logs', 'nivel', default='INFO')
        formato = self.config.get('logs', 'formato', 
                                  default='[%(asctime)s] %(levelname)s: %(message)s')
        
        logging.basicConfig(
            level=getattr(logging, nivel),
            format=formato,
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        
        logging.info("="*60)
        logging.info("Sistema iniciado")
        logging.info("="*60)


class BackupManager:
    """Gestor de backups automáticos"""
    
    def __init__(self, config: ConfigManager):
        self.config = config
        self.backup_dir = Path(config.get('archivos', 'directorio_backups', default='backups'))
        self.backup_dir.mkdir(exist_ok=True)
    
    def crear_backup(self, archivo_origen: str) -> bool:
        """Crea un backup del archivo Excel"""
        if not self.config.get('backups', 'automaticos', default=True):
            return True
        
        try:
            # Nombre del backup con timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            nombre_backup = f"backup_{timestamp}.xlsx"
            ruta_backup = self.backup_dir / nombre_backup
            
            # Copiar archivo
            shutil.copy2(archivo_origen, ruta_backup)
            logging.info(f"Backup creado: {nombre_backup}")
            
            # Limpiar backups antiguos
            self.limpiar_backups_antiguos()
            
            return True
        except Exception as e:
            logging.error(f"Error al crear backup: {e}")
            return False
    
    def limpiar_backups_antiguos(self):
        """Elimina backups antiguos según configuración"""
        max_backups = self.config.get('backups', 'max_backups', default=10)
        backups = sorted(self.backup_dir.glob('backup_*.xlsx'), reverse=True)
        
        # Eliminar backups excedentes
        for backup in backups[max_backups:]:
            try:
                backup.unlink()
                logging.info(f"Backup antiguo eliminado: {backup.name}")
            except Exception as e:
                logging.error(f"Error al eliminar backup {backup.name}: {e}")


class InterfazUI:
    """Clase para manejar la interfaz de usuario con colores"""
    
    def __init__(self, config: ConfigManager):
        self.config = config
        self.usar_colores = config.get('interfaz', 'usar_colores', default=True) and COLORAMA_AVAILABLE
    
    def limpiar_pantalla(self):
        """Limpia la consola"""
        if self.config.get('interfaz', 'limpiar_pantalla', default=True):
            os.system('clear' if os.name != 'nt' else 'cls')
    
    def print_header(self, texto: str, color=None):
        """Imprime un header estilizado"""
        if self.usar_colores and color:
            print(f"\n{color}{'='*60}")
            print(f"   {texto}")
            print(f"{'='*60}{Style.RESET_ALL}")
        else:
            print(f"\n{'='*60}")
            print(f"   {texto}")
            print(f"{'='*60}")
    
    def print_subheader(self, texto: str, color=None):
        """Imprime un subheader"""
        if self.usar_colores and color:
            print(f"\n{color}{'-'*60}")
            print(f"   {texto}")
            print(f"{'-'*60}{Style.RESET_ALL}")
        else:
            print(f"\n{'-'*60}")
            print(f"   {texto}")
            print(f"{'-'*60}")
    
    def print_exito(self, mensaje: str):
        """Imprime mensaje de éxito"""
        if self.usar_colores:
            print(f"\n{Fore.GREEN}✅ {mensaje}{Style.RESET_ALL}")
        else:
            print(f"\n✅ {mensaje}")
    
    def print_error(self, mensaje: str):
        """Imprime mensaje de error"""
        if self.usar_colores:
            print(f"\n{Fore.RED}❌ {mensaje}{Style.RESET_ALL}")
        else:
            print(f"\n❌ {mensaje}")
    
    def print_advertencia(self, mensaje: str):
        """Imprime mensaje de advertencia"""
        if self.usar_colores:
            print(f"\n{Fore.YELLOW}⚠️  {mensaje}{Style.RESET_ALL}")
        else:
            print(f"\n⚠️  {mensaje}")
    
    def print_info(self, mensaje: str):
        """Imprime mensaje informativo"""
        if self.usar_colores:
            print(f"{Fore.CYAN}{mensaje}{Style.RESET_ALL}")
        else:
            print(mensaje)


class SistemaCalificaciones:
    """Clase principal del sistema de calificaciones"""
    
    def __init__(self):
        self.config = ConfigManager()
        self.log_manager = LogManager(self.config)
        self.backup_manager = BackupManager(self.config)
        self.ui = InterfazUI(self.config)
        self.estudiantes = {}
        self.wb = None
        self.ws = None
        self.archivo_excel = self.config.get('archivos', 'excel_principal', default='grupo001.xlsx')
    
    def cargar_datos_excel(self) -> bool:
        """Carga los datos del archivo Excel"""
        try:
            self.wb = openpyxl.load_workbook(self.archivo_excel)
            self.ws = self.wb.active
            self.estudiantes = {}
            
            # Leer datos (empezando desde la fila 2 para saltar encabezados)
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay nombre
                    nombre = str(row[0]).strip().upper()
                    calificacion = row[1] if row[1] is not None else 0
                    min_aprobatoria = self.config.get('calificaciones', 'minima_aprobatoria', default=6.0)
                    self.estudiantes[nombre] = {
                        'calificacion': calificacion,
                        'estado': 'APROBADO' if calificacion >= min_aprobatoria else 'REPROBADO'
                    }
            
            logging.info(f"Datos cargados: {len(self.estudiantes)} estudiantes")
            return True
        except FileNotFoundError:
            self.ui.print_error(f"No se encontró el archivo {self.archivo_excel}")
            logging.error(f"Archivo no encontrado: {self.archivo_excel}")
            return False
        except Exception as e:
            self.ui.print_error(f"Error al cargar el archivo: {e}")
            logging.error(f"Error al cargar Excel: {e}")
            return False
    
    def guardar_cambios(self) -> bool:
        """Guarda cambios en el Excel con backup automático"""
        try:
            # Crear backup antes de guardar
            self.backup_manager.crear_backup(self.archivo_excel)
            
            # Guardar cambios
            self.wb.save(self.archivo_excel)
            logging.info("Cambios guardados exitosamente")
            return True
        except Exception as e:
            self.ui.print_error(f"Error al guardar: {e}")
            logging.error(f"Error al guardar Excel: {e}")
            return False
    
    def mostrar_menu_principal(self):
        """Muestra el menú principal de acceso"""
        self.ui.print_header("📚 SISTEMA DE CALIFICACIONES PRO 2.0 📚", Fore.CYAN + Style.BRIGHT)
        print("\n1. 👨‍🎓 Acceso para ESTUDIANTES")
        print("2. 👨‍🏫 Acceso para ADMINISTRADOR (Maestro)")
        print("3. 📊 Exportar Reportes")
        print("4. 🚪 Salir")
        print("\n" + "="*60)
    
    def mostrar_menu_admin(self):
        """Muestra el menú del administrador"""
        self.ui.print_header("👨‍🏫 PANEL DE ADMINISTRADOR 👨‍🏫", Fore.MAGENTA + Style.BRIGHT)
        print("\n1. 🔍 Consultar calificación de un estudiante")
        print("2. 📊 Ver todos los estudiantes")
        print("3. ➕ Agregar nuevo estudiante")
        print("4. ✏️  Modificar calificación")
        print("5. 🗑️  Eliminar estudiante")
        print("6. 📈 Estadísticas del grupo")
        print("7. 📁 Gestión de backups")
        print("8. 📜 Ver logs del sistema")
        print("9. 🔙 Volver al menú principal")
        print("\n" + "="*60)
    
    def mostrar_menu_estudiante(self):
        """Muestra el menú del estudiante"""
        self.ui.print_header("👨‍🎓 PANEL DE ESTUDIANTE 👨‍🎓", Fore.GREEN + Style.BRIGHT)
        print("\n1. 🔍 Consultar mi calificación")
        print("2. 📊 Ver estadísticas del grupo")
        print("3. 🔙 Volver al menú principal")
        print("\n" + "="*60)
    
    def consultar_estudiante(self, modo='admin'):
        """Consulta la calificación de un estudiante específico"""
        titulo = "🔍 CONSULTAR CALIFICACIÓN DE ESTUDIANTE" if modo == 'admin' else "🔍 CONSULTAR MI CALIFICACIÓN"
        self.ui.print_subheader(titulo, Fore.YELLOW)
        
        nombre = input("\nIngresa el nombre completo en MAYÚSCULAS (ej. JUAN PEREZ): ").strip().upper()
        
        if nombre in self.estudiantes:
            info = self.estudiantes[nombre]
            self.ui.print_exito("Estudiante encontrado!")
            print(f"\n   👤 Nombre: {nombre}")
            print(f"   📝 Calificación: {info['calificacion']}")
            print(f"   📊 Estado: {info['estado']}")
            
            if info['estado'] == 'APROBADO':
                self.ui.print_info("\n   🎉 ¡Felicidades! Has aprobado la materia.")
            else:
                self.ui.print_info("\n   😔 Lo siento, no has aprobado. ¡Sigue esforzándote!")
            
            logging.info(f"Consulta de estudiante: {nombre} - Calificación: {info['calificacion']}")
        else:
            self.ui.print_error(f"El estudiante '{nombre}' no se encuentra en el sistema.")
            if modo == 'estudiante':
                self.ui.print_info("   💡 Verifica que escribiste tu nombre correctamente en MAYÚSCULAS.")
            logging.warning(f"Estudiante no encontrado: {nombre}")
    
    def ver_todos_estudiantes(self):
        """Muestra todos los estudiantes y sus calificaciones"""
        self.ui.print_subheader("📊 LISTA COMPLETA DE ESTUDIANTES", Fore.YELLOW)
        
        if not self.estudiantes:
            self.ui.print_error("No hay estudiantes registrados.")
            return
        
        # Ordenar por nombre
        estudiantes_ordenados = sorted(self.estudiantes.items())
        
        print(f"\n{Fore.CYAN if self.ui.usar_colores else ''}{'NOMBRE':<30} {'CALIFICACIÓN':<15} {'ESTADO':<15}{Style.RESET_ALL if self.ui.usar_colores else ''}")
        print("-"*60)
        
        for nombre, info in estudiantes_ordenados:
            simbolo = "✅" if info['estado'] == 'APROBADO' else "❌"
            color = Fore.GREEN if info['estado'] == 'APROBADO' else Fore.RED
            if self.ui.usar_colores:
                print(f"{nombre:<30} {info['calificacion']:<15} {color}{simbolo} {info['estado']:<15}{Style.RESET_ALL}")
            else:
                print(f"{nombre:<30} {info['calificacion']:<15} {simbolo} {info['estado']:<15}")
        
        print("-"*60)
        print(f"Total de estudiantes: {len(self.estudiantes)}")
        logging.info("Vista de todos los estudiantes")
    
    def agregar_estudiante(self):
        """Agrega un nuevo estudiante al sistema"""
        self.ui.print_subheader("➕ AGREGAR NUEVO ESTUDIANTE", Fore.YELLOW)
        
        nombre = input("\nIngresa el nombre completo en MAYÚSCULAS: ").strip().upper()
        
        if nombre in self.estudiantes:
            self.ui.print_advertencia(f"El estudiante '{nombre}' ya existe en el sistema.")
            return
        
        try:
            calificacion = float(input("Ingresa la calificación (0-10): "))
            if calificacion < 0 or calificacion > 10:
                self.ui.print_error("La calificación debe estar entre 0 y 10.")
                return
        except ValueError:
            self.ui.print_error("Calificación inválida.")
            return
        
        # Agregar al diccionario
        min_aprobatoria = self.config.get('calificaciones', 'minima_aprobatoria', default=6.0)
        estado = 'APROBADO' if calificacion >= min_aprobatoria else 'REPROBADO'
        self.estudiantes[nombre] = {
            'calificacion': calificacion,
            'estado': estado
        }
        
        # Agregar al Excel
        nueva_fila = self.ws.max_row + 1
        self.ws[f'A{nueva_fila}'] = nombre
        self.ws[f'B{nueva_fila}'] = calificacion
        self.ws[f'C{nueva_fila}'] = f'=IF(B{nueva_fila}>={min_aprobatoria}, "Aprobado", "Reprobado")'
        
        if self.guardar_cambios():
            self.ui.print_exito(f"Estudiante '{nombre}' agregado exitosamente!")
            print(f"   📝 Calificación: {calificacion}")
            print(f"   📊 Estado: {estado}")
            logging.info(f"Estudiante agregado: {nombre} - Calificación: {calificacion}")
    
    def modificar_calificacion(self):
        """Modifica la calificación de un estudiante"""
        self.ui.print_subheader("✏️  MODIFICAR CALIFICACIÓN", Fore.YELLOW)
        
        nombre = input("\nIngresa el nombre del estudiante: ").strip().upper()
        
        if nombre not in self.estudiantes:
            self.ui.print_error(f"El estudiante '{nombre}' no se encuentra en el sistema.")
            return
        
        print(f"\nCalificación actual: {self.estudiantes[nombre]['calificacion']}")
        
        try:
            nueva_calificacion = float(input("Ingresa la nueva calificación (0-10): "))
            if nueva_calificacion < 0 or nueva_calificacion > 10:
                self.ui.print_error("La calificación debe estar entre 0 y 10.")
                return
        except ValueError:
            self.ui.print_error("Calificación inválida.")
            return
        
        # Actualizar diccionario
        min_aprobatoria = self.config.get('calificaciones', 'minima_aprobatoria', default=6.0)
        estado = 'APROBADO' if nueva_calificacion >= min_aprobatoria else 'REPROBADO'
        calificacion_anterior = self.estudiantes[nombre]['calificacion']
        self.estudiantes[nombre]['calificacion'] = nueva_calificacion
        self.estudiantes[nombre]['estado'] = estado
        
        # Actualizar Excel
        for row in range(2, self.ws.max_row + 1):
            if str(self.ws[f'A{row}'].value).strip().upper() == nombre:
                self.ws[f'B{row}'] = nueva_calificacion
                break
        
        if self.guardar_cambios():
            self.ui.print_exito("Calificación actualizada exitosamente!")
            print(f"   📝 Nueva calificación: {nueva_calificacion}")
            print(f"   📊 Estado: {estado}")
            logging.info(f"Calificación modificada: {nombre} - De {calificacion_anterior} a {nueva_calificacion}")
    
    def eliminar_estudiante(self):
        """Elimina un estudiante del sistema"""
        self.ui.print_subheader("🗑️  ELIMINAR ESTUDIANTE", Fore.RED)
        
        nombre = input("\nIngresa el nombre del estudiante a eliminar: ").strip().upper()
        
        if nombre not in self.estudiantes:
            self.ui.print_error(f"El estudiante '{nombre}' no se encuentra en el sistema.")
            return
        
        self.ui.print_advertencia(f"¿Estás seguro de eliminar a '{nombre}'?")
        confirmacion = input("Escribe 'SI' para confirmar: ").strip().upper()
        
        if confirmacion != 'SI':
            self.ui.print_error("Operación cancelada.")
            return
        
        # Eliminar del diccionario
        del self.estudiantes[nombre]
        
        # Eliminar del Excel
        fila_eliminar = None
        for row in range(2, self.ws.max_row + 1):
            if str(self.ws[f'A{row}'].value).strip().upper() == nombre:
                fila_eliminar = row
                break
        
        if fila_eliminar:
            self.ws.delete_rows(fila_eliminar, 1)
            if self.guardar_cambios():
                self.ui.print_exito(f"Estudiante '{nombre}' eliminado exitosamente.")
                logging.info(f"Estudiante eliminado: {nombre}")
    
    def mostrar_estadisticas(self):
        """Muestra estadísticas del grupo"""
        self.ui.print_subheader("📈 ESTADÍSTICAS DEL GRUPO", Fore.YELLOW)
        
        if not self.estudiantes:
            self.ui.print_error("No hay estudiantes registrados.")
            return
        
        calificaciones = [info['calificacion'] for info in self.estudiantes.values()]
        aprobados = sum(1 for info in self.estudiantes.values() if info['estado'] == 'APROBADO')
        reprobados = len(self.estudiantes) - aprobados
        
        promedio = sum(calificaciones) / len(calificaciones)
        maxima = max(calificaciones)
        minima = min(calificaciones)
        
        print(f"\n📊 Total de estudiantes: {len(self.estudiantes)}")
        print(f"✅ Aprobados: {aprobados} ({aprobados/len(self.estudiantes)*100:.1f}%)")
        print(f"❌ Reprobados: {reprobados} ({reprobados/len(self.estudiantes)*100:.1f}%)")
        print(f"\n📈 Calificación promedio: {promedio:.2f}")
        print(f"🏆 Calificación más alta: {maxima}")
        print(f"📉 Calificación más baja: {minima}")
        
        # Estudiante(s) con mejor calificación
        mejores = [nombre for nombre, info in self.estudiantes.items() if info['calificacion'] == maxima]
        print(f"\n🥇 Mejor(es) estudiante(s):")
        for nombre in mejores:
            print(f"   - {nombre}")
        
        logging.info("Estadísticas consultadas")
    
    def gestionar_backups(self):
        """Gestiona los backups del sistema"""
        self.ui.print_subheader("📁 GESTIÓN DE BACKUPS", Fore.YELLOW)
        
        backups = sorted(self.backup_manager.backup_dir.glob('backup_*.xlsx'), reverse=True)
        
        if not backups:
            self.ui.print_info("No hay backups disponibles.")
            return
        
        print(f"\n📦 Backups encontrados: {len(backups)}\n")
        for i, backup in enumerate(backups[:10], 1):  # Mostrar últimos 10
            tamaño = backup.stat().st_size / 1024  # KB
            fecha_mod = datetime.fromtimestamp(backup.stat().st_mtime)
            print(f"{i}. {backup.name} - {tamaño:.1f} KB - {fecha_mod.strftime('%d/%m/%Y %H:%M:%S')}")
        
        print(f"\n💡 Directorio de backups: {self.backup_manager.backup_dir}")
        logging.info("Gestión de backups consultada")
    
    def ver_logs(self):
        """Muestra los últimos logs del sistema"""
        self.ui.print_subheader("📜 LOGS DEL SISTEMA", Fore.YELLOW)
        
        log_file = self.log_manager.log_dir / f"sistema_{datetime.now().strftime('%Y%m%d')}.log"
        
        if not log_file.exists():
            self.ui.print_info("No hay logs disponibles para hoy.")
            return
        
        try:
            with open(log_file, 'r', encoding='utf-8') as f:
                lineas = f.readlines()
                ultimas_lineas = lineas[-20:]  # Últimas 20 líneas
                
                print(f"\n📄 Últimas {len(ultimas_lineas)} entradas del log:\n")
                for linea in ultimas_lineas:
                    print(linea.rstrip())
                
                print(f"\n💡 Archivo de log: {log_file}")
        except Exception as e:
            self.ui.print_error(f"Error al leer logs: {e}")
    
    def exportar_reporte_csv(self):
        """Exporta un reporte en formato CSV"""
        self.ui.print_subheader("📊 EXPORTAR REPORTE CSV", Fore.YELLOW)
        
        if not self.estudiantes:
            self.ui.print_error("No hay estudiantes para exportar.")
            return
        
        try:
            # Crear directorio de reportes
            reportes_dir = Path(self.config.get('archivos', 'directorio_reportes', default='reportes'))
            reportes_dir.mkdir(exist_ok=True)
            
            # Nombre del archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archivo_csv = reportes_dir / f"reporte_{timestamp}.csv"
            
            # Escribir CSV
            with open(archivo_csv, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['NOMBRE', 'CALIFICACIÓN', 'ESTADO'])
                
                for nombre, info in sorted(self.estudiantes.items()):
                    writer.writerow([nombre, info['calificacion'], info['estado']])
            
            self.ui.print_exito(f"Reporte exportado: {archivo_csv.name}")
            print(f"   📁 Ubicación: {archivo_csv}")
            logging.info(f"Reporte CSV exportado: {archivo_csv}")
            
        except Exception as e:
            self.ui.print_error(f"Error al exportar reporte: {e}")
            logging.error(f"Error al exportar CSV: {e}")
    
    def verificar_admin(self) -> bool:
        """Verifica la contraseña del administrador"""
        self.ui.print_subheader("🔐 ACCESO DE ADMINISTRADOR", Fore.MAGENTA)
        
        password_correcta = self.config.get('seguridad', 'admin_password', default='admin123')
        max_intentos = self.config.get('seguridad', 'max_intentos_login', default=3)
        intentos = max_intentos
        
        while intentos > 0:
            password = getpass.getpass(f"\nIngresa la contraseña (intentos restantes: {intentos}): ")
            
            if password == password_correcta:
                self.ui.print_exito("Acceso concedido. Bienvenido, Maestro.")
                logging.info("Acceso de administrador concedido")
                time.sleep(1)
                return True
            else:
                intentos -= 1
                if intentos > 0:
                    self.ui.print_error(f"Contraseña incorrecta. Te quedan {intentos} intentos.")
                    logging.warning(f"Intento de acceso fallido. Intentos restantes: {intentos}")
                else:
                    self.ui.print_error("Acceso denegado. Demasiados intentos fallidos.")
                    logging.warning("Acceso de administrador denegado - Demasiados intentos")
                    time.sleep(2)
        
        return False
    
    def panel_estudiante(self):
        """Panel de acceso para estudiantes"""
        while True:
            self.mostrar_menu_estudiante()
            opcion = input("\nSelecciona una opción (1-3): ").strip()
            
            if opcion == '1':
                self.consultar_estudiante(modo='estudiante')
            elif opcion == '2':
                self.mostrar_estadisticas()
            elif opcion == '3':
                self.ui.print_info("\n🔙 Volviendo al menú principal...")
                time.sleep(1)
                break
            else:
                self.ui.print_error("Opción inválida. Por favor selecciona 1-3.")
            
            input("\n⏎ Presiona ENTER para continuar...")
            self.ui.limpiar_pantalla()
    
    def panel_admin(self):
        """Panel de acceso para administradores"""
        if not self.verificar_admin():
            return
        
        self.ui.limpiar_pantalla()
        
        while True:
            self.mostrar_menu_admin()
            opcion = input("\nSelecciona una opción (1-9): ").strip()
            
            if opcion == '1':
                self.consultar_estudiante(modo='admin')
            elif opcion == '2':
                self.ver_todos_estudiantes()
            elif opcion == '3':
                self.agregar_estudiante()
            elif opcion == '4':
                self.modificar_calificacion()
            elif opcion == '5':
                self.eliminar_estudiante()
            elif opcion == '6':
                self.mostrar_estadisticas()
            elif opcion == '7':
                self.gestionar_backups()
            elif opcion == '8':
                self.ver_logs()
            elif opcion == '9':
                self.ui.print_info("\n🔙 Volviendo al menú principal...")
                time.sleep(1)
                break
            else:
                self.ui.print_error("Opción inválida. Por favor selecciona 1-9.")
            
            input("\n⏎ Presiona ENTER para continuar...")
            self.ui.limpiar_pantalla()
    
    def menu_exportar_reportes(self):
        """Menú de exportación de reportes"""
        self.ui.print_header("📊 EXPORTAR REPORTES", Fore.BLUE + Style.BRIGHT)
        print("\n1. 📄 Exportar a CSV")
        print("2. 📈 Exportar estadísticas completas")
        print("3. 🔙 Volver")
        print("\n" + "="*60)
        
        opcion = input("\nSelecciona una opción (1-3): ").strip()
        
        if opcion == '1':
            self.exportar_reporte_csv()
        elif opcion == '2':
            self.exportar_estadisticas_completas()
        elif opcion == '3':
            return
        else:
            self.ui.print_error("Opción inválida.")
        
        input("\n⏎ Presiona ENTER para continuar...")
    
    def exportar_estadisticas_completas(self):
        """Exporta estadísticas completas a un archivo de texto"""
        self.ui.print_subheader("📈 EXPORTAR ESTADÍSTICAS COMPLETAS", Fore.YELLOW)
        
        if not self.estudiantes:
            self.ui.print_error("No hay estudiantes para generar estadísticas.")
            return
        
        try:
            reportes_dir = Path(self.config.get('archivos', 'directorio_reportes', default='reportes'))
            reportes_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            archivo_txt = reportes_dir / f"estadisticas_{timestamp}.txt"
            
            with open(archivo_txt, 'w', encoding='utf-8') as f:
                f.write("="*70 + "\n")
                f.write("  REPORTE DE ESTADÍSTICAS - SISTEMA DE CALIFICACIONES PRO 2.0\n")
                f.write("="*70 + "\n\n")
                f.write(f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n\n")
                
                # Estadísticas generales
                calificaciones = [info['calificacion'] for info in self.estudiantes.values()]
                aprobados = sum(1 for info in self.estudiantes.values() if info['estado'] == 'APROBADO')
                reprobados = len(self.estudiantes) - aprobados
                promedio = sum(calificaciones) / len(calificaciones)
                maxima = max(calificaciones)
                minima = min(calificaciones)
                
                f.write("ESTADÍSTICAS GENERALES\n")
                f.write("-" * 70 + "\n")
                f.write(f"Total de estudiantes: {len(self.estudiantes)}\n")
                f.write(f"Aprobados: {aprobados} ({aprobados/len(self.estudiantes)*100:.1f}%)\n")
                f.write(f"Reprobados: {reprobados} ({reprobados/len(self.estudiantes)*100:.1f}%)\n")
                f.write(f"Calificación promedio: {promedio:.2f}\n")
                f.write(f"Calificación más alta: {maxima}\n")
                f.write(f"Calificación más baja: {minima}\n\n")
                
                # Lista de estudiantes
                f.write("LISTA COMPLETA DE ESTUDIANTES\n")
                f.write("-" * 70 + "\n")
                f.write(f"{'NOMBRE':<35} {'CALIFICACIÓN':<15} {'ESTADO':<15}\n")
                f.write("-" * 70 + "\n")
                
                for nombre, info in sorted(self.estudiantes.items()):
                    f.write(f"{nombre:<35} {info['calificacion']:<15} {info['estado']:<15}\n")
                
                f.write("\n" + "="*70 + "\n")
                f.write("Fin del reporte\n")
            
            self.ui.print_exito(f"Estadísticas exportadas: {archivo_txt.name}")
            print(f"   📁 Ubicación: {archivo_txt}")
            logging.info(f"Estadísticas completas exportadas: {archivo_txt}")
            
        except Exception as e:
            self.ui.print_error(f"Error al exportar estadísticas: {e}")
            logging.error(f"Error al exportar estadísticas: {e}")
    
    def ejecutar(self):
        """Ejecuta el sistema principal"""
        self.ui.limpiar_pantalla()
        self.ui.print_header("🎓 SISTEMA DE CALIFICACIONES PRO 2.0 🎓", Fore.CYAN + Style.BRIGHT)
        time.sleep(1)
        
        # Cargar datos
        print("\n⏳ Cargando datos...")
        if not self.cargar_datos_excel():
            self.ui.print_error("No se pudo iniciar el sistema.")
            return
        
        self.ui.print_exito(f"Datos cargados: {len(self.estudiantes)} estudiantes encontrados.")
        time.sleep(1)
        self.ui.limpiar_pantalla()
        
        while True:
            self.mostrar_menu_principal()
            opcion = input("\nSelecciona una opción (1-4): ").strip()
            
            if opcion == '1':
                self.ui.limpiar_pantalla()
                self.ui.print_info("\n👨‍🎓 Bienvenido, Estudiante")
                time.sleep(1)
                self.ui.limpiar_pantalla()
                self.panel_estudiante()
                self.ui.limpiar_pantalla()
            elif opcion == '2':
                self.ui.limpiar_pantalla()
                self.panel_admin()
                self.ui.limpiar_pantalla()
            elif opcion == '3':
                self.ui.limpiar_pantalla()
                self.menu_exportar_reportes()
                self.ui.limpiar_pantalla()
            elif opcion == '4':
                self.ui.print_info("\n👋 ¡Gracias por usar el sistema! Hasta pronto.")
                logging.info("Sistema cerrado")
                time.sleep(1)
                break
            else:
                self.ui.print_error("Opción inválida. Por favor selecciona 1-4.")
                time.sleep(1)
                self.ui.limpiar_pantalla()


def main():
    """Función principal"""
    try:
        sistema = SistemaCalificaciones()
        sistema.ejecutar()
    except KeyboardInterrupt:
        print("\n\n⚠️  Sistema interrumpido por el usuario.")
        logging.info("Sistema interrumpido por el usuario")
    except Exception as e:
        print(f"\n❌ Error crítico: {e}")
        logging.critical(f"Error crítico: {e}")


if __name__ == "__main__":
    main()

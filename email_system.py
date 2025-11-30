"""
Sistema de Notificaciones por Email - Sistema de Calificaciones
Envía calificaciones y números de cuenta a los estudiantes por correo electrónico
"""

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from datetime import datetime
import time

# ============================================================================
# CONFIGURACIÓN DE EMAIL
# ============================================================================

# IMPORTANTE: Para usar Gmail necesitas:
# 1. Habilitar "Verificación en 2 pasos" en tu cuenta de Google
# 2. Crear una "Contraseña de aplicación" específica para este sistema
# 3. Ir a: https://myaccount.google.com/apppasswords

EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'puerto': 587,
    'remitente_email': 'tu_email@gmail.com',  # TU EMAIL AQUÍ
    'remitente_password': 'tu_contraseña_de_aplicacion',  # CONTRASEÑA DE APLICACIÓN AQUÍ
    'remitente_nombre': 'Sistema de Calificaciones'
}


class GestorEmails:
    """Clase para gestionar el envío de emails"""
    
    def __init__(self, config=EMAIL_CONFIG):
        self.config = config
        self.smtp_server = config['smtp_server']
        self.puerto = config['puerto']
        self.remitente_email = config['remitente_email']
        self.remitente_password = config['remitente_password']
        self.remitente_nombre = config['remitente_nombre']
    
    def verificar_configuracion(self):
        """Verifica que la configuración de email esté completa"""
        if 'tu_email@gmail.com' in self.remitente_email or 'tu_contraseña' in self.remitente_password:
            print("\n⚠️  CONFIGURACIÓN INCOMPLETA")
            print("━" * 60)
            print("\nPara enviar emails, debes configurar:")
            print("\n1. Tu email en EMAIL_CONFIG['remitente_email']")
            print("2. Tu contraseña de aplicación en EMAIL_CONFIG['remitente_password']")
            print("\n📖 Instrucciones:")
            print("   1. Ve a: https://myaccount.google.com/apppasswords")
            print("   2. Crea una contraseña de aplicación")
            print("   3. Cópiala en el archivo email_system.py")
            print("\n━" * 60)
            return False
        return True
    
    def enviar_email(self, destinatario, asunto, cuerpo_html, cuerpo_texto=None):
        """
        Envía un email
        
        Args:
            destinatario (str): Email del destinatario
            asunto (str): Asunto del email
            cuerpo_html (str): Contenido del email en HTML
            cuerpo_texto (str): Contenido alternativo en texto plano
        
        Returns:
            bool: True si se envió correctamente, False si hubo error
        """
        try:
            # Crear mensaje
            mensaje = MIMEMultipart('alternative')
            mensaje['From'] = f"{self.remitente_nombre} <{self.remitente_email}>"
            mensaje['To'] = destinatario
            mensaje['Subject'] = asunto
            mensaje['Date'] = datetime.now().strftime('%a, %d %b %Y %H:%M:%S %z')
            
            # Agregar cuerpo en texto plano (fallback)
            if cuerpo_texto:
                parte_texto = MIMEText(cuerpo_texto, 'plain', 'utf-8')
                mensaje.attach(parte_texto)
            
            # Agregar cuerpo en HTML
            parte_html = MIMEText(cuerpo_html, 'html', 'utf-8')
            mensaje.attach(parte_html)
            
            # Conectar y enviar
            with smtplib.SMTP(self.smtp_server, self.puerto) as servidor:
                servidor.starttls()  # Seguridad TLS
                servidor.login(self.remitente_email, self.remitente_password)
                servidor.send_message(mensaje)
            
            return True
            
        except Exception as e:
            print(f"❌ Error al enviar email a {destinatario}: {e}")
            return False
    
    def generar_email_calificacion(self, nombre, calificacion, estado, numero_cuenta):
        """Genera el contenido HTML del email de calificación"""
        
        # Color según el estado
        color_estado = "#4CAF50" if estado == "APROBADO" else "#f44336"
        emoji_estado = "🎉" if estado == "APROBADO" else "📚"
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                }}
                .container {{
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    padding: 30px;
                    border-radius: 10px;
                    box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                }}
                .content {{
                    background: white;
                    padding: 30px;
                    border-radius: 8px;
                    margin-top: 20px;
                }}
                h1 {{
                    color: white;
                    margin: 0;
                    font-size: 28px;
                    text-align: center;
                }}
                .info-box {{
                    background: #f5f5f5;
                    padding: 20px;
                    border-radius: 8px;
                    margin: 20px 0;
                }}
                .info-item {{
                    display: flex;
                    justify-content: space-between;
                    padding: 10px 0;
                    border-bottom: 1px solid #e0e0e0;
                }}
                .info-item:last-child {{
                    border-bottom: none;
                }}
                .label {{
                    font-weight: bold;
                    color: #555;
                }}
                .value {{
                    color: #333;
                }}
                .estado {{
                    display: inline-block;
                    padding: 8px 16px;
                    border-radius: 20px;
                    font-weight: bold;
                    color: white;
                    background: {color_estado};
                    text-align: center;
                    font-size: 18px;
                }}
                .numero-cuenta {{
                    background: #2196F3;
                    color: white;
                    padding: 15px;
                    border-radius: 8px;
                    text-align: center;
                    font-size: 24px;
                    font-weight: bold;
                    letter-spacing: 2px;
                    margin: 20px 0;
                }}
                .footer {{
                    text-align: center;
                    color: #777;
                    font-size: 12px;
                    margin-top: 30px;
                    padding-top: 20px;
                    border-top: 1px solid #e0e0e0;
                }}
                .warning {{
                    background: #fff3cd;
                    border-left: 4px solid #ffc107;
                    padding: 15px;
                    margin: 20px 0;
                    border-radius: 4px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>📚 Sistema de Calificaciones</h1>
                
                <div class="content">
                    <h2>Hola {nombre},</h2>
                    <p>Te enviamos la información de tu calificación y credenciales de acceso al sistema.</p>
                    
                    <div class="info-box">
                        <div class="info-item">
                            <span class="label">👤 Estudiante:</span>
                            <span class="value">{nombre}</span>
                        </div>
                        <div class="info-item">
                            <span class="label">📝 Calificación:</span>
                            <span class="value" style="font-size: 24px; font-weight: bold; color: {color_estado};">{calificacion}</span>
                        </div>
                        <div class="info-item">
                            <span class="label">📊 Estado:</span>
                            <span class="value"><span class="estado">{emoji_estado} {estado}</span></span>
                        </div>
                    </div>
                    
                    <div class="warning">
                        <strong>🔐 Tu Número de Cuenta (CONFIDENCIAL)</strong>
                        <div class="numero-cuenta">{numero_cuenta}</div>
                        <p style="margin: 10px 0 0 0; font-size: 14px;">
                            <strong>⚠️ IMPORTANTE:</strong> Este número es personal e intransferible. 
                            Úsalo para consultar tu calificación en el sistema.
                        </p>
                    </div>
                    
                    <div style="background: #e3f2fd; padding: 15px; border-radius: 8px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #1976d2;">📱 Cómo acceder al sistema:</h3>
                        <ol style="margin: 10px 0; padding-left: 20px;">
                            <li>Ejecuta el sistema de calificaciones</li>
                            <li>Selecciona "Acceso para ESTUDIANTES"</li>
                            <li>Ingresa tu número de cuenta: <strong>{numero_cuenta}</strong></li>
                            <li>Consulta tu información</li>
                        </ol>
                    </div>
                    
                    {"<div style='background: #c8e6c9; padding: 15px; border-radius: 8px; text-align: center;'><strong>🎉 ¡Felicidades por aprobar!</strong><br>Continúa con ese excelente desempeño.</div>" if estado == "APROBADO" else "<div style='background: #ffccbc; padding: 15px; border-radius: 8px; text-align: center;'><strong>📚 No te desanimes</strong><br>Sigue esforzándote, ¡tú puedes lograrlo!</div>"}
                    
                    <div class="footer">
                        <p><strong>Sistema de Calificaciones Académico</strong></p>
                        <p>Este es un mensaje automático. Por favor no respondas a este correo.</p>
                        <p style="color: #999; font-size: 11px;">
                            Enviado el {datetime.now().strftime('%d de %B de %Y a las %H:%M')}
                        </p>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Versión texto plano (fallback)
        texto = f"""
        Sistema de Calificaciones
        
        Hola {nombre},
        
        Te enviamos la información de tu calificación:
        
        - Estudiante: {nombre}
        - Calificación: {calificacion}
        - Estado: {estado}
        - Número de Cuenta: {numero_cuenta}
        
        ⚠️ IMPORTANTE: Tu número de cuenta es confidencial. Úsalo para consultar 
        tu calificación en el sistema.
        
        Cómo acceder:
        1. Ejecuta el sistema de calificaciones
        2. Selecciona "Acceso para ESTUDIANTES"
        3. Ingresa tu número de cuenta: {numero_cuenta}
        
        ---
        Sistema de Calificaciones Académico
        {datetime.now().strftime('%d de %B de %Y')}
        """
        
        return html, texto
    
    def enviar_calificaciones_masivas(self, archivo_excel='grupo001.xlsx'):
        """
        Envía calificaciones a todos los estudiantes que tengan email registrado
        
        Args:
            archivo_excel (str): Ruta del archivo Excel con los datos
        
        Returns:
            dict: Estadísticas del envío
        """
        if not self.verificar_configuracion():
            return None
        
        print("\n📧 ENVÍO MASIVO DE CALIFICACIONES")
        print("━" * 60)
        
        try:
            # Cargar Excel
            wb = openpyxl.load_workbook(archivo_excel)
            ws = wb.active
            
            # Verificar columnas
            headers = [cell.value for cell in ws[1]]
            
            # Buscar índices de columnas
            idx_nombre = headers.index('NOMBRE DE ALUMNO') if 'NOMBRE DE ALUMNO' in headers else 0
            idx_calificacion = headers.index('CALIFICACION') if 'CALIFICACION' in headers else 1
            idx_numero_cuenta = headers.index('NUMERO DE CUENTA') if 'NUMERO DE CUENTA' in headers else 3
            idx_email = headers.index('EMAIL') if 'EMAIL' in headers else None
            
            if idx_email is None:
                print("\n⚠️  No se encontró la columna 'EMAIL' en el archivo Excel")
                print("   Agrega una columna 'EMAIL' con los correos de los estudiantes")
                return None
            
            # Estadísticas
            total = 0
            enviados = 0
            fallidos = 0
            sin_email = 0
            
            # Procesar cada estudiante
            print("\n⏳ Enviando emails...")
            print()
            
            for row in range(2, ws.max_row + 1):
                nombre = ws.cell(row, idx_nombre + 1).value
                calificacion = ws.cell(row, idx_calificacion + 1).value
                numero_cuenta = ws.cell(row, idx_numero_cuenta + 1).value
                email = ws.cell(row, idx_email + 1).value
                
                if not nombre:
                    continue
                
                total += 1
                nombre = str(nombre).strip()
                
                # Verificar si tiene email
                if not email or str(email).strip() == '' or '@' not in str(email):
                    print(f"⊘  {nombre:<30} - Sin email registrado")
                    sin_email += 1
                    continue
                
                email = str(email).strip()
                
                # Determinar estado
                estado = 'APROBADO' if calificacion >= 6 else 'REPROBADO'
                
                # Generar contenido del email
                html, texto = self.generar_email_calificacion(
                    nombre, calificacion, estado, numero_cuenta
                )
                
                # Enviar email
                asunto = f"📚 Tu Calificación y Número de Cuenta - {nombre}"
                
                if self.enviar_email(email, asunto, html, texto):
                    print(f"✅ {nombre:<30} → {email}")
                    enviados += 1
                else:
                    print(f"❌ {nombre:<30} → {email} (Error)")
                    fallidos += 1
                
                # Pausa para no saturar el servidor
                time.sleep(1)
            
            # Resumen
            print("\n" + "━" * 60)
            print("📊 RESUMEN DEL ENVÍO")
            print("━" * 60)
            print(f"Total de estudiantes: {total}")
            print(f"✅ Enviados exitosamente: {enviados}")
            print(f"❌ Fallidos: {fallidos}")
            print(f"⊘  Sin email: {sin_email}")
            print("━" * 60)
            
            return {
                'total': total,
                'enviados': enviados,
                'fallidos': fallidos,
                'sin_email': sin_email
            }
            
        except Exception as e:
            print(f"\n❌ Error en el envío masivo: {e}")
            return None
    
    def enviar_email_individual(self, email, nombre, calificacion, estado, numero_cuenta):
        """Envía un email individual a un estudiante"""
        
        if not self.verificar_configuracion():
            return False
        
        html, texto = self.generar_email_calificacion(nombre, calificacion, estado, numero_cuenta)
        asunto = f"📚 Tu Calificación y Número de Cuenta - {nombre}"
        
        return self.enviar_email(email, asunto, html, texto)


def agregar_columna_email():
    """Agrega la columna EMAIL al archivo Excel si no existe"""
    try:
        archivo = 'grupo001.xlsx'
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
        
        # Verificar si ya existe
        headers = [cell.value for cell in ws[1]]
        if 'EMAIL' in headers:
            print("✅ La columna EMAIL ya existe")
            return
        
        # Agregar columna EMAIL en la columna E (5)
        ws.cell(1, 5).value = 'EMAIL'
        wb.save(archivo)
        
        print("✅ Columna EMAIL agregada al archivo Excel")
        print("💡 Ahora puedes agregar los emails de los estudiantes en esa columna")
        
    except Exception as e:
        print(f"❌ Error: {e}")


def menu_principal():
    """Menú principal del sistema de emails"""
    gestor = GestorEmails()
    
    while True:
        print("\n" + "="*60)
        print("   📧 SISTEMA DE NOTIFICACIONES POR EMAIL 📧")
        print("="*60)
        print("\n1. 📤 Enviar calificaciones a todos los estudiantes")
        print("2. 📧 Enviar email a un estudiante específico")
        print("3. ⚙️  Agregar columna EMAIL al Excel")
        print("4. 🔧 Verificar configuración")
        print("5. 🚪 Salir")
        print("\n" + "="*60)
        
        opcion = input("\nSelecciona una opción (1-5): ").strip()
        
        if opcion == '1':
            print("\n⚠️  Se enviará un email a TODOS los estudiantes con email registrado")
            confirmacion = input("¿Deseas continuar? (SI/NO): ").strip().upper()
            if confirmacion == 'SI':
                gestor.enviar_calificaciones_masivas()
            else:
                print("❌ Operación cancelada")
        
        elif opcion == '2':
            print("\n📧 ENVIAR EMAIL INDIVIDUAL")
            print("-" * 60)
            email = input("Email del estudiante: ").strip()
            nombre = input("Nombre del estudiante: ").strip().upper()
            try:
                calificacion = float(input("Calificación: "))
                numero_cuenta = input("Número de cuenta: ").strip()
                estado = 'APROBADO' if calificacion >= 6 else 'REPROBADO'
                
                if gestor.enviar_email_individual(email, nombre, calificacion, estado, numero_cuenta):
                    print(f"\n✅ Email enviado exitosamente a {email}")
                else:
                    print(f"\n❌ Error al enviar email")
            except ValueError:
                print("❌ Calificación inválida")
        
        elif opcion == '3':
            agregar_columna_email()
        
        elif opcion == '4':
            print("\n🔧 VERIFICACIÓN DE CONFIGURACIÓN")
            print("-" * 60)
            if gestor.verificar_configuracion():
                print("\n✅ Configuración completa")
                print(f"   Servidor SMTP: {gestor.smtp_server}")
                print(f"   Puerto: {gestor.puerto}")
                print(f"   Email remitente: {gestor.remitente_email}")
            else:
                print("\n❌ Configuración incompleta. Revisa las instrucciones arriba.")
        
        elif opcion == '5':
            print("\n👋 ¡Hasta pronto!")
            break
        
        else:
            print("\n❌ Opción inválida")
        
        input("\n⏎ Presiona ENTER para continuar...")


if __name__ == "__main__":
    menu_principal()
